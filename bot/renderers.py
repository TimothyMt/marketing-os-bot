"""
Output renderers — convert parsed agent output to deliverable formats.

Per-skill primary_deliverable determines which format is sent as main attachment:
  - HTML: all skills support (general purpose, mobile-safe)
  - EXCEL: content_calendar (table grid), performance_audit (data tables)
  - MARKDOWN: ad_copy, video_scripts, briefs (deliverable for downstream tools)

HTML always generated as fallback. Excel/Markdown generated when primary_deliverable matches.
"""
import io
import re
import logging
from datetime import datetime
from typing import Optional

from agents.skills import OutputFormat, PrimaryDeliverable

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────
# Parsers — handle 3 output format variants
# ─────────────────────────────────────────────────────────────────

def parse_strategic_output(text: str) -> dict:
    """Parse STRATEGIC_4_SECTION output into {insight, summary, benchmarks, detail}."""
    from bot.html_report import parse_agent_output
    return parse_agent_output(text)


def parse_operational_deliverable(text: str) -> dict:
    """Parse OPERATIONAL_DELIVERABLE output into {summary, deliverable, raw}."""
    # ALWAYS preserve raw để Excel renderer có thể fallback
    result = {"summary": "", "deliverable": "", "raw": text}

    # Match "## 🎯 Tóm tắt nhanh" section
    summary_match = re.search(
        r"##\s*🎯[^\n]*\n+(.*?)(?=\n##\s|\Z)",
        text, flags=re.DOTALL
    )
    if summary_match:
        result["summary"] = summary_match.group(1).strip()

    # Match "## 📄 Deliverable" section
    deliverable_match = re.search(
        r"##\s*📄[^\n]*\n+(.*?)(?=\Z)",
        text, flags=re.DOTALL
    )
    if deliverable_match:
        result["deliverable"] = deliverable_match.group(1).strip()

    # Fallback: nếu parse fail, dùng cả text làm deliverable
    if not result["summary"] and not result["deliverable"]:
        result["deliverable"] = text.strip()

    return result


def parse_operational_analysis(text: str) -> dict:
    """Parse OPERATIONAL_ANALYSIS output into {summary, kpi_table, root_cause, actions, forecast}."""
    result = {"summary": "", "kpi_table": "", "root_cause": "", "actions": "", "forecast": "", "raw": text}

    patterns = {
        "summary":    r"##\s*📊\s*Tóm tắt[^\n]*\n+(.*?)(?=\n##\s|\Z)",
        "kpi_table":  r"##\s*📈[^\n]*\n+(.*?)(?=\n##\s|\Z)",
        "root_cause": r"##\s*🔬[^\n]*\n+(.*?)(?=\n##\s|\Z)",
        "actions":    r"##\s*🎯[^\n]*\n+(.*?)(?=\n##\s|\Z)",
        "forecast":   r"##\s*📉[^\n]*\n+(.*?)(?=\Z)",
    }
    for key, pat in patterns.items():
        m = re.search(pat, text, flags=re.DOTALL)
        if m:
            result[key] = m.group(1).strip()

    if not any(v for k, v in result.items() if k != "raw"):
        result["summary"] = text.strip()[:2000]

    return result


def parse_by_format(text: str, output_format: OutputFormat) -> dict:
    """Dispatch parser based on output format."""
    if output_format == OutputFormat.STRATEGIC_4_SECTION:
        return parse_strategic_output(text)
    elif output_format == OutputFormat.OPERATIONAL_DELIVERABLE:
        return parse_operational_deliverable(text)
    elif output_format == OutputFormat.OPERATIONAL_ANALYSIS:
        return parse_operational_analysis(text)
    return {"raw": text}


# ─────────────────────────────────────────────────────────────────
# Telegram card formatters
# ─────────────────────────────────────────────────────────────────

def format_telegram_card(
    skill_name: str,
    skill_label: str,
    skill_emoji: str,
    parsed: dict,
    output_format: OutputFormat,
    file_attached_hint: Optional[str] = None,
) -> str:
    """Build Telegram preview card. Long content always goes to HTML/file."""
    header = f"*{skill_emoji} {skill_label.upper()}*"
    separator = "━" * 25
    parts = [header, separator, ""]

    if output_format == OutputFormat.STRATEGIC_4_SECTION:
        if parsed.get("insight"):
            insight = parsed["insight"].strip().strip('"').strip("'")
            parts.append("💡 *Insight quan trọng nhất:*")
            parts.append(f"_{insight}_")
            parts.append("")
        if parsed.get("summary"):
            parts.append("📌 *Tóm tắt:*")
            parts.append(parsed["summary"].strip())
            parts.append("")
        if parsed.get("benchmarks"):
            parts.append("📊 *Benchmarks:*")
            parts.append(parsed["benchmarks"].strip())
            parts.append("")

    elif output_format == OutputFormat.OPERATIONAL_DELIVERABLE:
        if parsed.get("summary"):
            parts.append("🎯 *Tóm tắt nhanh:*")
            parts.append(parsed["summary"].strip())
            parts.append("")

    elif output_format == OutputFormat.OPERATIONAL_ANALYSIS:
        if parsed.get("summary"):
            parts.append("📊 *Tổng quan:*")
            parts.append(parsed["summary"].strip())
            parts.append("")

    # Hint about attached file
    if file_attached_hint:
        parts.append(f"📎 _{file_attached_hint}_")
    else:
        parts.append("📎 _Xem chi tiết trong file đính kèm bên dưới_")

    return "\n".join(parts)


# ─────────────────────────────────────────────────────────────────
# Markdown deliverable file generator
# ─────────────────────────────────────────────────────────────────

def render_markdown_file(
    skill_name: str,
    skill_label: str,
    parsed: dict,
    output_format: OutputFormat,
    business_name: str = "",
) -> bytes:
    """Render skill output as a .md file for download (designer/dev/creator workflow)."""
    lines = [
        f"# {skill_label} — {business_name or 'Marketing OS'}",
        f"*Generated by Max — AI CMO · {datetime.now().strftime('%d/%m/%Y · %H:%M')}*",
        "",
        "---",
        "",
    ]

    if output_format == OutputFormat.OPERATIONAL_DELIVERABLE:
        if parsed.get("summary"):
            lines += ["## 🎯 Tóm tắt nhanh", "", parsed["summary"].strip(), "", "---", ""]
        if parsed.get("deliverable"):
            lines += [parsed["deliverable"].strip()]
    elif output_format == OutputFormat.STRATEGIC_4_SECTION:
        if parsed.get("insight"):
            lines += [f"> 💡 **Insight:** {parsed['insight'].strip().strip(chr(34)).strip(chr(39))}", ""]
        if parsed.get("summary"):
            lines += ["## 🎯 Tóm tắt", "", parsed["summary"].strip(), ""]
        if parsed.get("benchmarks"):
            lines += ["## 📊 Benchmarks", "", parsed["benchmarks"].strip(), ""]
        if parsed.get("detail"):
            lines += ["## 📄 Phân tích chi tiết", "", parsed["detail"].strip(), ""]
    elif output_format == OutputFormat.OPERATIONAL_ANALYSIS:
        # Dump all sections
        for key in ["summary", "kpi_table", "root_cause", "actions", "forecast"]:
            if parsed.get(key):
                lines += [f"## {key.replace('_', ' ').title()}", "", parsed[key].strip(), "", "---", ""]
    else:
        lines += [parsed.get("raw", "")]

    return "\n".join(lines).encode("utf-8")


# ─────────────────────────────────────────────────────────────────
# Excel renderer — for content_calendar + performance_audit
# ─────────────────────────────────────────────────────────────────

def _clean_cell(value) -> str:
    """Strip markdown chars (** __ *) khỏi cell content. Trả về str."""
    if value is None:
        return ""
    s = str(value)
    # Remove bold/italic markers
    s = re.sub(r"\*\*(.+?)\*\*", r"\1", s)
    s = re.sub(r"__(.+?)__", r"\1", s)
    s = re.sub(r"(?<!\*)\*(?!\*)([^*\n]+?)\*(?!\*)", r"\1", s)
    s = re.sub(r"(?<!_)_(?!_)([^_\n]+?)_(?!_)", r"\1", s)
    return s.strip()


def _safe_sheet_name(raw: str, idx: int, used: set) -> str:
    """Sheet name an toàn (≤31 chars, không có : \\ / ? * [ ], unique)."""
    cleaned = re.sub(r"[:\\/?*\[\]]", " ", raw or "")
    # Remove emoji prefix to save chars
    cleaned = re.sub(r"^[^\w\d]+", "", cleaned).strip()
    cleaned = cleaned[:28].strip() or f"Sheet {idx+1}"
    # Ensure unique
    base = cleaned
    n = 2
    while cleaned in used:
        suffix = f" {n}"
        cleaned = base[:28 - len(suffix)] + suffix
        n += 1
    used.add(cleaned)
    return cleaned


def _is_keyvalue_table(headers: list, rows: list[list]) -> bool:
    """Detect mini key-value table: 2 columns, headers look generic, repeating field names."""
    if len(headers) != 2:
        return False
    GENERIC = {"field", "value", "key", "thông tin", "chi tiết", "metadata"}
    h0 = (headers[0] or "").lower().strip("* ")
    h1 = (headers[1] or "").lower().strip("* ")
    if h0 in GENERIC or h1 in GENERIC:
        return True
    # If first column values look like field names (Ngày, Kênh, Pillar...)
    FIELD_HINTS = {"ngày", "kênh", "pillar", "funnel", "source", "format", "hook",
                   "cta", "angle", "topic", "tier", "platform", "campaign"}
    first_col_vals = [_clean_cell(r[0]).lower() for r in rows[:5]]
    if sum(1 for v in first_col_vals if any(h in v for h in FIELD_HINTS)) >= 2:
        return True
    return False


def _pivot_keyvalue_tables(tables: list[tuple]) -> Optional[tuple]:
    """Convert nhiều mini key-value tables thành 1 master sheet.
    Mỗi mini-table thành 1 row, key = column header.
    Returns (title, headers, rows) hoặc None nếu không có cụm KV nào.
    """
    kv_tables = [(t, h, r) for (t, h, r) in tables if _is_keyvalue_table(h, r)]
    if len(kv_tables) < 2:
        return None

    # Collect all unique field names across all KV tables (preserve order from first)
    all_fields: list[str] = []
    seen_fields = set()
    for _, _, rows in kv_tables:
        for row in rows:
            if not row: continue
            field = _clean_cell(row[0])
            field_lower = field.lower()
            if field and field_lower not in seen_fields:
                seen_fields.add(field_lower)
                all_fields.append(field)

    # Build master rows: 1 row per table
    master_headers = ["Bài"] + all_fields
    master_rows = []
    for tbl_title, _, rows in kv_tables:
        # Strip emoji + clean title
        clean_title = _clean_cell(tbl_title or "")
        clean_title = re.sub(r"^[^\w\d]+", "", clean_title).strip()[:60]
        row_dict = {_clean_cell(r[0]).lower(): _clean_cell(r[1] if len(r) > 1 else "") for r in rows if r}
        master_rows.append([clean_title] + [row_dict.get(f.lower(), "") for f in all_fields])

    return ("📊 Tổng hợp", master_headers, master_rows)


def _split_table_by_week(headers: list, rows: list[list]) -> Optional[dict]:
    """Nếu table có cột 'Tuần' → group rows theo tuần.
    Returns {week_label: rows} hoặc None nếu không có cột Tuần.
    Dùng cho content_generator output để tạo sheet riêng cho từng tuần.
    """
    if not headers or not rows:
        return None
    # Tìm index cột 'Tuần' (case-insensitive)
    week_idx = None
    for i, h in enumerate(headers):
        h_clean = (h or "").strip().lower().lstrip("*").rstrip("*").strip()
        if h_clean in ("tuần", "tuan", "week", "tuần ", "wk"):
            week_idx = i
            break
    if week_idx is None:
        return None

    groups: dict[str, list] = {}
    for row in rows:
        if len(row) <= week_idx:
            continue
        week_label = _clean_cell(row[week_idx]).strip() or "Khác"
        # Normalize "Tuần 1" / "Tuan 1" / "Week 1"
        week_label = re.sub(r"(?i)^(tu[aâà]n|week|wk)\s*", "Tuần ", week_label).strip()
        groups.setdefault(week_label, []).append(row)

    if len(groups) < 2:
        return None  # Chỉ có 1 tuần, không cần split
    return groups


def render_excel_file(
    skill_name: str,
    skill_label: str,
    parsed: dict,
    output_format: OutputFormat,
    business_name: str = "",
) -> Optional[bytes]:
    """Render skill output as .xlsx.
    Special handling cho content_generator: split rows by 'Tuần' column.
    Default: auto-pivot mini key-value tables thành 1 master overview.

    Layout:
    - content_generator: Sheet "Tổng hợp" + sheet riêng cho mỗi Tuần
    - Default: Sheet "Tổng hợp" + non-KV tables
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.warning("openpyxl not installed — falling back to no Excel export")
        return None

    if output_format == OutputFormat.OPERATIONAL_DELIVERABLE:
        # Include raw để cover case LLM output table ngoài section "Deliverable"
        full_text = "\n\n".join(filter(None, [
            parsed.get("deliverable", ""),
            parsed.get("summary", ""),
            parsed.get("raw", ""),
        ]))
    elif output_format == OutputFormat.OPERATIONAL_ANALYSIS:
        full_text = "\n\n".join(
            parsed.get(k, "") for k in ["summary", "kpi_table", "root_cause", "actions", "forecast"]
        )
        # Fallback to raw nếu structured parse fail
        if not full_text.strip():
            full_text = parsed.get("raw", "")
    else:
        full_text = parsed.get("detail", "") + "\n\n" + parsed.get("raw", "")

    tables = _extract_markdown_tables(full_text)
    if not tables:
        logger.warning("render_excel_file [%s]: no markdown tables found (full_text len=%d)",
                       skill_name, len(full_text))
        return None

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name="Arial")
    body_align = Alignment(vertical="top", wrap_text=True)
    title_font = Font(bold=True, size=14, name="Arial")

    # Detect & merge key-value tables
    overview = _pivot_keyvalue_tables(tables)
    if overview:
        sheets_to_render = [overview]
        # Add non-KV tables
        for t in tables:
            if not _is_keyvalue_table(t[1], t[2]):
                sheets_to_render.append(t)
    else:
        sheets_to_render = tables[:8]  # cap 8 sheets

    # SPECIAL — Content Generator: chỉ giữ MASTER table (có cột Tuần + Bài),
    # bỏ qua các mini-table phụ trong content (size guides, comparison tables, etc.)
    if skill_name == "content_generator":
        master_table = None
        for t_title, t_headers, t_rows in tables:
            cleaned_headers_lower = [_clean_cell(h).lower().strip() for h in t_headers]
            has_tuan = any(h in ("tuần", "tuan", "week") for h in cleaned_headers_lower)
            has_bai = any(h in ("bài", "bai", "post", "#") for h in cleaned_headers_lower)
            if has_tuan and has_bai:
                master_table = (t_title, t_headers, t_rows)
                break

        if master_table:
            s_title, s_headers, s_rows = master_table
            sheets_to_render = []
            # Overview sheet
            sheets_to_render.append((f"📊 Tổng hợp ({len(s_rows)} bài)", s_headers, s_rows))
            # Split by week
            week_groups = _split_table_by_week(s_headers, s_rows)
            if week_groups:
                for week_label in sorted(week_groups.keys(), key=lambda x: (len(x), x)):
                    week_rows = week_groups[week_label]
                    sheets_to_render.append((f"{week_label} ({len(week_rows)} bài)", s_headers, week_rows))
        else:
            # Không có master table → LLM output thiếu. Vẫn render những gì có để debug
            logger.warning("content_generator: master table (Tuần+Bài columns) not found in output. Falling back to default render.")

    used_names = set()
    for idx, (table_title, headers, rows) in enumerate(sheets_to_render):
        raw_name = table_title or f"Bảng {idx+1}"
        sheet_name = _safe_sheet_name(raw_name, idx, used_names)
        ws = wb.create_sheet(title=sheet_name)

        # Title row
        clean_title = _clean_cell(table_title or "")
        if clean_title:
            ws.append([clean_title])
            ws["A1"].font = title_font
            ws.append([])

        # Headers
        clean_headers = [_clean_cell(h) for h in headers]
        header_row = ws.max_row + 1 if ws.max_row > 0 else 1
        ws.append(clean_headers)
        for col_idx in range(1, len(clean_headers) + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Data rows — strip markdown
        for row in rows:
            cleaned_row = [_clean_cell(c) for c in row]
            ws.append(cleaned_row)
            r_idx = ws.max_row
            for c_idx in range(1, len(cleaned_row) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = body_font
                cell.alignment = body_align

        # Auto column width (capped 60)
        for col_idx in range(1, len(clean_headers) + 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            max_len = max(
                len(clean_headers[col_idx-1]) if col_idx-1 < len(clean_headers) else 0,
                *[len(str(r[col_idx-1] or "")) for r in rows if col_idx-1 < len(r)],
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        # Freeze header row
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _extract_markdown_tables(text: str) -> list[tuple[str, list[str], list[list[str]]]]:
    """Extract markdown tables from text. Returns list of (title, headers, rows).
    Title is the nearest preceding heading (###/####)."""
    tables = []
    lines = text.split("\n")

    # State: 'looking_for_table' | 'in_header' | 'in_body'
    current_title = ""
    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # Track title (last seen heading)
        h_match = re.match(r"^#{1,5}\s+(.+)$", line)
        if h_match:
            current_title = h_match.group(1).strip()

        # Detect table header row: | col | col |
        if re.match(r"^\|.+\|$", line) and i + 1 < len(lines):
            sep_line = lines[i + 1].strip()
            # Separator row: |---|---|
            if re.match(r"^\|[\s:|-]+\|$", sep_line):
                headers = [c.strip() for c in line.strip("|").split("|")]
                rows = []
                j = i + 2
                while j < len(lines) and re.match(r"^\|.+\|$", lines[j].strip()):
                    row_cells = [c.strip() for c in lines[j].strip().strip("|").split("|")]
                    # Pad to match header length
                    while len(row_cells) < len(headers):
                        row_cells.append("")
                    rows.append(row_cells[:len(headers)])
                    j += 1
                if rows:
                    tables.append((current_title, headers, rows))
                i = j
                continue
        i += 1

    return tables
