"""
Funnel Mapper — map ToFu/MoFu/BoFu per channel từ Campaign Brief.

Input:  Campaign dict (output của campaign_intake)
Output: Funnel Map list → feed vào content_calendar + content_generator
"""
from __future__ import annotations

import io
import json
import logging
import re
from typing import Optional

from storage.models import Session

logger = logging.getLogger(__name__)

_JSON_ARRAY_RE = re.compile(r"```(?:json)?\s*(\[.*?\])\s*```", re.DOTALL)


# ─────────────────────────────────────────────────────────────────
# Parse
# ─────────────────────────────────────────────────────────────────

def parse_funnel_map(text: str) -> Optional[list]:
    """Extract funnel map JSON array từ LLM output. Pure."""
    if not text:
        return None
    candidates = _JSON_ARRAY_RE.findall(text)
    if not candidates:
        stripped = text.strip()
        if stripped.startswith("[") and stripped.endswith("]"):
            candidates = [stripped]
    for raw in candidates:
        try:
            obj = json.loads(raw)
            if isinstance(obj, list) and all(
                isinstance(item, dict) and "channel" in item for item in obj
            ):
                return obj
        except (json.JSONDecodeError, ValueError):
            continue
    return None


# ─────────────────────────────────────────────────────────────────
# Generator
# ─────────────────────────────────────────────────────────────────

async def generate_funnel_map(session: Session, campaign: dict) -> list:
    """Generate ToFu/MoFu/BoFu mapping per channel. 1 LLM call (Sonnet)."""
    from tools.llm_router import call as router_call, TaskType, AllProvidersFailedError
    from agents.campaign_intake_prompts import FUNNEL_MAPPER_SYSTEM, build_funnel_mapper_user
    from agents.campaign_scope_library import format_scope_for_prompt

    industry       = session.profile.industry or ""
    industry_scope = format_scope_for_prompt(industry) if industry else ""
    user_msg       = build_funnel_mapper_user(campaign=campaign, industry_scope=industry_scope)

    try:
        result = await router_call(
            task_type  = TaskType.GENERIC_CREATIVE,
            system     = FUNNEL_MAPPER_SYSTEM,
            user       = user_msg,
            max_tokens = 5000,
        )
        raw        = result.get("output", "")
        funnel_map = parse_funnel_map(raw)
        if funnel_map:
            return funnel_map
        logger.warning("funnel_map parse failed, raw[:300]: %s", raw[:300])
    except AllProvidersFailedError as e:
        logger.error("generate_funnel_map failed: %s", e)

    return _fallback_funnel_map(
        channels  = campaign.get("channels") or [],
        objective = campaign.get("objective", "mix"),
    )


def _fallback_funnel_map(channels: list, objective: str) -> list:
    """Generic fallback khi LLM fail."""
    ratio_map = {
        "awareness":  "60/30/10",
        "branding":   "50/40/10",
        "conversion": "30/30/40",
        "mix":        "50/30/20",
    }
    ratio = ratio_map.get(objective, "50/30/20")
    result = []
    for ch in channels:
        result.append({
            "channel": ch,
            "ratio": ratio,
            "tofu": {
                "goal":           "Tiếp cận tệp mới chưa biết brand",
                "formats":        ["Short video", "Image post"],
                "content_angles": ["Giáo dục", "Giải trí"],
                "cta":            "Follow / Xem thêm",
                "volume":         "3/tuần",
            },
            "mofu": {
                "goal":           "Build trust với người đã biết brand",
                "formats":        ["Testimonial", "Behind-the-scenes"],
                "content_angles": ["Social proof", "Chuyên môn"],
                "cta":            "Comment / Lưu lại",
                "volume":         "2/tuần",
            },
            "bofu": {
                "goal":           "Convert người đang cân nhắc",
                "formats":        ["Offer reveal", "CTA trực tiếp"],
                "content_angles": ["Urgency", "Value proof"],
                "cta":            "Mua ngay / Inbox / Book",
                "volume":         "1/tuần",
            },
            "calendar_note": "Bám tỷ lệ ToFu/MoFu/BoFu theo objective campaign",
        })
    return result


# ─────────────────────────────────────────────────────────────────
# Card renderer
# ─────────────────────────────────────────────────────────────────

def render_funnel_map_card(funnel_map: list) -> str:
    """Format funnel map → Telegram card. Pure."""
    if not funnel_map:
        return "_(Không có funnel map)_"

    STAGE_EMOJI = {"tofu": "🔵", "mofu": "🟡", "bofu": "🟢"}
    STAGE_LABEL = {"tofu": "ToFu", "mofu": "MoFu", "bofu": "BoFu"}

    lines = ["🗺 *Funnel Map — chiến lược từng kênh*", ""]

    for ch_map in funnel_map:
        ch    = ch_map.get("channel", "")
        ratio = ch_map.get("ratio", "")
        lines.append(f"📡 *{ch}* _(tỷ lệ {ratio})_")

        for stage in ("tofu", "mofu", "bofu"):
            s       = ch_map.get(stage) or {}
            emoji   = STAGE_EMOJI[stage]
            label   = STAGE_LABEL[stage]
            formats = ", ".join(s.get("formats") or [])
            vol     = s.get("volume", "")
            lines.append(f"  {emoji} *{label}* ({vol}): {s.get('goal', '')}")
            lines.append(f"     Format: {formats}")
            lines.append(f"     CTA: _{s.get('cta', '')}_")

        note = ch_map.get("calendar_note", "")
        if note:
            lines.append(f"  💡 _{note}_")
        lines.append("")

    lines.append("_Sếp duyệt funnel map để em build Content Calendar nhé?_")
    return "\n".join(lines).strip()


def render_funnel_map_summary(funnel_map: list) -> str:
    """Tóm tắt NGẮN cho Telegram — chi tiết đầy đủ nằm trong file HTML."""
    if not funnel_map:
        return "_(Không có funnel map)_"

    lines = ["🗺 *Funnel Map — tóm tắt theo kênh*", ""]
    for ch_map in funnel_map:
        ch    = ch_map.get("channel", "")
        ratio = ch_map.get("ratio", "")
        lines.append(f"📡 *{ch}* — _ToFu/MoFu/BoFu {ratio}_")
    lines.append("")
    lines.append("_📄 Chi tiết format · content angle · CTA · volume từng kênh → xem file HTML đính kèm._")
    return "\n".join(lines)


def build_funnel_map_markdown(funnel_map: list) -> str:
    """Markdown ĐẦY ĐỦ cho HTML report — mọi stage, mọi kênh."""
    if not funnel_map:
        return "_(Không có funnel map)_"

    STAGE_LABEL = {"tofu": "🔵 ToFu", "mofu": "🟡 MoFu", "bofu": "🟢 BoFu"}
    parts = []
    for ch_map in funnel_map:
        ch    = ch_map.get("channel", "")
        ratio = ch_map.get("ratio", "")
        parts.append(f"### 📡 {ch} — tỷ lệ ToFu/MoFu/BoFu {ratio}")
        for stage in ("tofu", "mofu", "bofu"):
            s = ch_map.get(stage) or {}
            formats = ", ".join(s.get("formats") or [])
            angles  = ", ".join(s.get("content_angles") or [])
            parts.append(f"**{STAGE_LABEL[stage]}** ({s.get('volume', '')}) — {s.get('goal', '')}")
            parts.append(f"- Format: {formats}")
            parts.append(f"- Content angles: {angles}")
            parts.append(f"- CTA: {s.get('cta', '')}")
        note = ch_map.get("calendar_note", "")
        if note:
            parts.append(f"> 💡 Lưu ý calendar: {note}")
        parts.append("")
    return "\n".join(parts).strip()


# ─────────────────────────────────────────────────────────────────
# Calendar bridge
# ─────────────────────────────────────────────────────────────────

def funnel_map_to_calendar_input(funnel_map: list, campaign: dict) -> dict:
    """Prepare funnel_map + campaign → structured input cho content_calendar skill.

    Đây là "interface" giữa FunnelMapper và ContentCalendar —
    content_calendar nhận dict này thay vì raw campaign.
    """
    channel_plans = []
    for ch_map in funnel_map:
        ch = ch_map.get("channel", "")
        channel_plans.append({
            "channel": ch,
            "ratio":   ch_map.get("ratio", "50/30/20"),
            "weekly_volume": {
                stage: (ch_map.get(stage) or {}).get("volume", "2/tuần")
                for stage in ("tofu", "mofu", "bofu")
            },
            "formats": {
                stage: (ch_map.get(stage) or {}).get("formats", [])
                for stage in ("tofu", "mofu", "bofu")
            },
            "content_angles": {
                stage: (ch_map.get(stage) or {}).get("content_angles", [])
                for stage in ("tofu", "mofu", "bofu")
            },
            "ctas": {
                stage: (ch_map.get(stage) or {}).get("cta", "")
                for stage in ("tofu", "mofu", "bofu")
            },
            "calendar_note": ch_map.get("calendar_note", ""),
        })

    return {
        "campaign_name":    campaign.get("name", ""),
        "objective":        campaign.get("objective", "mix"),
        "objective_detail": campaign.get("objective_detail", ""),
        "duration_days":    campaign.get("duration_days", 30),
        "audience":         campaign.get("audience", ""),
        "content_pillars":  campaign.get("content_pillars", []),
        "brand_voice":      campaign.get("brand_voice", {}),
        "budget_total":     campaign.get("budget_total", ""),
        "kpi_targets":      campaign.get("kpi_targets", []),
        "extra_notes":      campaign.get("extra_notes", ""),
        "channel_plans":    channel_plans,
    }


# ─────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────

def build_funnel_map_excel(funnel_map: list, campaign_name: str = "") -> bytes:
    """Export full funnel map → .xlsx bytes (openpyxl).

    Sheet "Funnel Map": one row per (channel × stage).
    Columns: Kênh | Tỷ lệ | Giai đoạn | Mục tiêu | Formats | Content Angles | CTA | Volume
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    STAGE_LABELS = {"tofu": "ToFu 🔵", "mofu": "MoFu 🟡", "bofu": "BoFu 🟢"}
    STAGE_FILLS  = {
        "tofu": PatternFill("solid", fgColor="BDD7EE"),
        "mofu": PatternFill("solid", fgColor="FFE699"),
        "bofu": PatternFill("solid", fgColor="C6EFCE"),
    }
    HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
    BOLD         = Font(bold=True)
    WRAP         = Alignment(wrap_text=True, vertical="top")
    CENTER_WRAP  = Alignment(wrap_text=True, vertical="top", horizontal="center")
    THIN         = Side(style="thin", color="BFBFBF")
    BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Funnel Map"

    # ── Title row ────────────────────────────────────────────────
    title_text = f"Funnel Map — {campaign_name}" if campaign_name else "Funnel Map"
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font  = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Header row ───────────────────────────────────────────────
    HEADERS = ["Kênh", "Tỷ lệ\nToFu/MoFu/BoFu", "Giai đoạn", "Mục tiêu", "Formats", "Content Angles", "CTA", "Volume"]
    COL_WIDTHS = [20, 18, 14, 36, 32, 32, 26, 14]
    for col_idx, (hdr, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=2, column=col_idx, value=hdr)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER_WRAP
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 32

    # ── Data rows ────────────────────────────────────────────────
    current_row = 3
    for ch_map in funnel_map:
        ch    = ch_map.get("channel", "")
        ratio = ch_map.get("ratio", "")
        stage_rows_start = current_row

        for stage in ("tofu", "mofu", "bofu"):
            s = ch_map.get(stage) or {}
            formats = "\n".join(f"• {f}" for f in (s.get("formats") or []))
            angles  = "\n".join(f"• {a}" for a in (s.get("content_angles") or []))
            row_data = [
                ch,
                ratio,
                STAGE_LABELS[stage],
                s.get("goal", ""),
                formats,
                angles,
                s.get("cta", ""),
                s.get("volume", ""),
            ]
            fill = STAGE_FILLS[stage]
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.alignment = WRAP
                cell.border    = BORDER
                # Stage column gets stage colour; others get light row colour
                if col_idx == 3:
                    cell.fill = fill
                    cell.font = BOLD
                elif col_idx in (1, 2):
                    cell.font = BOLD
            ws.row_dimensions[current_row].height = max(
                40, 15 * max(1, len((s.get("formats") or [])), len((s.get("content_angles") or [])))
            )
            current_row += 1

        # Merge Kênh + Tỷ lệ columns across the 3 stage rows
        if current_row - stage_rows_start > 1:
            for merge_col in (1, 2):
                ws.merge_cells(
                    start_row=stage_rows_start, start_column=merge_col,
                    end_row=current_row - 1, end_column=merge_col,
                )
                merged = ws.cell(row=stage_rows_start, column=merge_col)
                merged.alignment = CENTER_WRAP
                merged.font      = BOLD

        # Calendar note row (span all cols)
        note = ch_map.get("calendar_note", "")
        if note:
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row,   end_column=8,
            )
            note_cell = ws.cell(row=current_row, column=1, value=f"💡 {note}")
            note_cell.font      = Font(italic=True, color="595959")
            note_cell.alignment = WRAP
            note_cell.border    = BORDER
            ws.row_dimensions[current_row].height = 20
            current_row += 1

    # ── Freeze header ────────────────────────────────────────────
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
